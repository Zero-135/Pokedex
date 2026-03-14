import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import subprocess

# ===============================
# CONFIGURACIÓN
# ===============================
carpeta = r"C:\Users\Walter Rivas\Documents\Imagenes\Pokemon\Icons"
archivo_excel = "nombres_archivos.xlsx"

# ===============================
# CREAR EXCEL
# ===============================
wb = Workbook()
ws = wb.active
ws.title = "Archivos"

# Encabezado
ws["A1"] = "Nombre de archivo"

fila = 2
for nombre in os.listdir(carpeta):
    ruta_completa = os.path.join(carpeta, nombre)
    if os.path.isfile(ruta_completa):
        ws[f"A{fila}"] = nombre.replace(".png", "")
        fila += 1

# ===============================
# AUTOAJUSTAR COLUMNA
# ===============================
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

# ===============================
# ACTIVAR FILTROS
# ===============================
ws.auto_filter.ref = ws.dimensions

# ===============================
# GUARDAR Y ABRIR
# ===============================
wb.save(archivo_excel)
subprocess.Popen([archivo_excel], shell=True)

print("Excel creado y abierto correctamente.")
