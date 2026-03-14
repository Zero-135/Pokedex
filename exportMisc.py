import os
import re
import sys
import platform
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ==============================
# CONFIGURACIÓN
# ==============================

wordFind = "types"
folder_path = r"C:\Users\Walter Rivas\Documents\PokeProject V3\PBS"
output_excel = wordFind + "_consolidado.xlsx"

# ==============================
# PROCESAMIENTO
# ==============================

all_data = []
all_keys = set()

for filename in os.listdir(folder_path):
    if filename.lower().startswith(wordFind) and filename.lower().endswith(".txt"):
        filepath = os.path.join(folder_path, filename)

        with open(filepath, "r", encoding="utf-8") as file:
            lines = file.readlines()

        current_entry = {}
        current_id = None

        for line in lines:
            line = line.strip()

            if not line or line.startswith("#"):
                continue

            id_match = re.match(r"\[(.*?)\]", line)
            if id_match:
                if current_id:
                    current_entry["Archivo"] = filename
                    current_entry["ID"] = current_id
                    all_data.append(current_entry)
                    all_keys.update(current_entry.keys())

                current_id = id_match.group(1)
                current_entry = {}

            elif "=" in line:
                key, value = line.split("=", 1)
                current_entry[key.strip()] = value.strip()

        if current_id:
            current_entry["Archivo"] = filename
            current_entry["ID"] = current_id
            all_data.append(current_entry)
            all_keys.update(current_entry.keys())

if not all_data:
    print("No se encontraron archivos " + wordFind + "*.txt")
    sys.exit()

# ==============================
# ORDENAR POR ID
# ==============================

all_data.sort(key=lambda x: x.get("ID", "").lower())

# ==============================
# CREAR EXCEL
# ==============================

wb = Workbook()
ws = wb.active
ws.title = wordFind

headers = ["Archivo", "ID"] + sorted(
    k for k in all_keys if k not in ["Archivo", "ID"]
)

ws.append(headers)

# 🔥 CABECERA EN NEGRITA
for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).font = Font(bold=True)

# Insertar datos
for entry in all_data:
    row = [entry.get(col, "") for col in headers]
    ws.append(row)

# ==============================
# FILTROS
# ==============================

ws.auto_filter.ref = ws.dimensions

# ==============================
# AUTOAJUSTAR COLUMNAS
# ==============================

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)

    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    ws.column_dimensions[col_letter].width = max_length + 2

# ==============================
# GUARDAR
# ==============================

wb.save(output_excel)
print("Archivo generado:", output_excel)

# ==============================
# ABRIR AUTOMÁTICAMENTE
# ==============================

if platform.system() == "Windows":
    os.startfile(output_excel)
elif platform.system() == "Darwin":
    subprocess.call(["open", output_excel])
else:
    subprocess.call(["xdg-open", output_excel])

sys.exit()